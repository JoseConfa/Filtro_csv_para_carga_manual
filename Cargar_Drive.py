"""
Cargar_Drive.py
===============
Módulo para crear y formatear archivos de Google Sheets.

Este módulo se encarga de:
- Crear nuevos archivos de Google Sheets
- Subir datos procesados a Google Sheets
- Aplicar formato profesional a las hojas
- Crear múltiples hojas dentro del mismo archivo

Funcionalidades principales:
- Creación automática de archivos con fecha
- Formato de datos para visualización óptima
- Manejo de múltiples DataFrames en un solo archivo
- Aplicación de bordes y formato profesional

Dependencias:
- openpyxl: Para manejo de formatos de Excel
- gspread_formatting: Para aplicar estilos a Google Sheets
- pandas: Para manipulación de datos

Autor: Sistema de Gestión de Pedidos
"""

from openpyxl.utils import get_column_letter
from gspread_formatting import *
import FiltroArgentina as f
import pandas as pd
from datetime import datetime

def cargar_excel(gc, archivo_final, archivo_andreani):
    """
    Crea un archivo de Google Sheets con los datos procesados.
    
    Esta función toma los DataFrames procesados y crea un archivo de Google Sheets
    con dos hojas: una para los datos principales y otra para el formato Andreani.
    
    Args:
        gc: Cliente autorizado de gspread
        archivo_final: DataFrame con datos procesados de Argentina
        archivo_andreani: DataFrame con datos procesados para Andreani
    
    Returns:
        None
    
    Proceso:
        1. Convierte DataFrames a strings para evitar errores de formato
        2. Extrae fecha del archivo para el nombre
        3. Crea archivo de Google Sheets con fecha
        4. Sube datos a la hoja principal
        5. Crea segunda hoja para datos de Andreani  
        6. Aplica formato profesional a ambas hojas
    """
    # Convertir DataFrame a strings para evitar problemas de formato
    dataframe1 = archivo_final.astype(str)

    # Extraer fecha del archivo para crear nombre descriptivo
    Fecha = None
    for index, row in dataframe1.iterrows():
        try:
            fecha_candidata = row['Created at']
            if pd.notna(fecha_candidata):
                fecha_dt = pd.to_datetime(fecha_candidata, errors='raise')
                Fecha = fecha_dt.strftime('%d-%m-%Y')
                break
        except:
            continue
    
    # Si no se encuentra fecha válida, usar fecha actual
    if Fecha is None:
        Fecha = datetime.now().strftime('%d-%m-%Y')

    # Crear nuevo archivo de Google Sheets con nombre descriptivo
    sh = gc.create('Archivo ' + Fecha + ' ARG')
    worksheet = sh.worksheet("Hoja 1")

    # Obtener dimensiones del DataFrame principal
    filas1, columnas1 = dataframe1.shape

    # Calcular rango de celdas usando notación Excel (A1:Z100, etc.)
    col_inicial1 = get_column_letter(1)
    col_final1 = get_column_letter(columnas1)
    range_str1 = f"{col_inicial1}1:{col_final1}{filas1}"

    # Convertir DataFrame a lista de listas para Google Sheets API
    data1 = dataframe1.values.tolist()

    # Limpiar valores "nan" que pueden aparecer por pandas
    valor_a_borrar = "nan"
    valor_a_reemplazar = ""
    for sublist in data1:
        for i in range(len(sublist)):
            if sublist[i] == valor_a_borrar:
                sublist[i] = valor_a_reemplazar

    # Subir datos a la hoja principal usando modo RAW para mantener formato original
    worksheet.update(range_str1, data1, value_input_option='RAW')

    # === PROCESAMIENTO DE LA SEGUNDA HOJA (ANDREANI) ===
    
    # Preparar datos del segundo DataFrame (Andreani)
    dataframe2 = archivo_andreani.astype(str)

    # Obtener dimensiones del DataFrame de Andreani
    filas2, columnas2 = dataframe2.shape

    # Crear segunda hoja específica para formato Andreani
    nueva_hoja = sh.add_worksheet(title="Andreani", rows=str(filas2), cols=str(columnas2))

    # Calcular rango de celdas para la segunda hoja
    col_inicial2 = get_column_letter(1)
    col_final2 = get_column_letter(columnas2)
    range_str2 = f"{col_inicial2}1:{col_final2}{filas2}"

    # Convertir segundo DataFrame a lista de listas
    data2 = dataframe2.values.tolist()

    # Limpiar valores "nan" en los datos de Andreani
    for sublist in data2:
        for i in range(len(sublist)):
            if sublist[i] == valor_a_borrar:
                sublist[i] = valor_a_reemplazar

    # Limpiar valores "n" sueltos que pueden aparecer en campos vacíos
    valor_a_borrar2 = "n"
    valor_a_reemplazar2 = ""
    for sublist in data2:
        for i in range(len(sublist)):
            if sublist[i] == valor_a_borrar2:
                sublist[i] = valor_a_reemplazar2

    # Subir datos a la hoja de Andreani
    nueva_hoja.update(range_str2, data2, value_input_option='RAW')

    # === APLICACIÓN DE FORMATO PROFESIONAL ===
    
    # Definir formato con bordes y texto en negrita para encabezados
    cell_format = CellFormat(
        textFormat=TextFormat(bold=True),
        borders=Borders(
            top=Border(style='SOLID'),
            bottom=Border(style='SOLID'),
            left=Border(style='SOLID'),
            right=Border(style='SOLID'),
        )
    )
    
    # Aplicar formato a ambas hojas para presentación profesional
    format_cell_range(worksheet, range_str1, cell_format)
    format_cell_range(nueva_hoja, range_str2, cell_format)
    format_cell_range(nueva_hoja, range_str2, cell_format)
